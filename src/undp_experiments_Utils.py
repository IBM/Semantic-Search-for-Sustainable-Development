"""
   Copyright 2018 IBM Corporation

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.
"""

import shelve
import parseundp
from PyPDF2 import PdfFileWriter, PdfFileReader
from difflib import SequenceMatcher as sm
from collections import OrderedDict
import os

from custom_model import CustomParVec
from custom_logger import CustomLogger
from pathlib import Path
from typing import Any

logger = CustomLogger(__name__, write_local=True)


#Loading the ground truth
#First we need a function to load our previous RIA data. This represents the ground truth as to what sentences match certain targets. 
#If we are testing with a country which we have a completed RIA for, we exclude that RIA. We also have functionality to only 
#load the target descriptions.

def loadTruth(template_data_path, exclude_ria=[], targets=None):
    '''
    If the ground truth matches aren't already saved, extract the ground truth from all prior RIA template2 and 
    create the target matches dictionary. Save this for future use. If a certain RIA is to be excluded, create
    a new dictionary that won't be saved.

    Args:
        template_data_path (string) : Directory to the RIA templates documents.
        exclude_ria (list[string])  : List of files to be excluded from RIA data extraction.
        targets (dict)              : Target matches dictionary of just target descriptions.

    Returns:
        target_matches (dict) : Dictionary of ground truth matches for each target.
    '''
    if targets:
        target_matches = targets
    elif len(exclude_ria) > 0:
        development_matches = parseundp.extract_template_data(template_data_path, exclude_ria)
        target_matches = parseundp.create_target_dictionary(development_matches)
    else:
        try:
            shelf = shelve.open('undp')
            target_matches = shelf['targets']
            shelf.close()
        
        # Do too many things in except block is not a good practice.
        except Exception as e:
            logger.error(e)
            shelf.close()
            development_matches = parseundp.extract_template_data(template_data_path)
            target_matches = parseundp.create_target_dictionary(development_matches)
            shelf = shelve.open('undp')
            shelf['targets'] = target_matches
            shelf.close()
            
    return target_matches


def getTargetDoc(template_data_path, exclude_ria=[]):
    '''
    Append the results of all prior RIAs to the corresponding target description.

    Args:
        template_data_path (string) : Directory to the RIA templates documents.
        exclude_ria (list[string])  : List of files to be excluded from RIA data extraction.
        targets (dict)              : Target matches dictionary of just target descriptions.

    Returns:
        target_matches (dict) : Dictionary of ground truth matches appended to its corresponding target description.
                                The dictionary value is a list of a single item.
    '''
    prior_ria_matches = loadTruth(template_data_path, exclude_ria)
    target_documents = {}
    for key in prior_ria_matches:
        doc = ''
        for val in prior_ria_matches[key]:
            if isinstance(val, str):
                doc += val
                doc += ' '
        target_documents[key] = [doc]
    return target_documents


def getInfo(par_vec, target_matches, data=False):
    '''
    If the ground truth matches aren't already saved, extract the ground truth from all prior RIA template2 and 
    create the target matches dictionary. Save this for future use. If a certain RIA is to be excluded, create
    a new dictionary that won't be saved.

    Args:
        par_vec (CustomParVec) : Embedding model to be used.
        target_matches (dict)  : Our target matches dictionary of.
        data (bool)    : Specify if we are only using target descriptions.

    Returns:
        targs (dict), targ_vecs(list), sents(list) : dictionary of sentences and target they match, list of embedded
                                                     ground truth vectors, list of sentences
    '''
    targs = {}
    targ_vecs = []
    sents = []
    for key, val in target_matches.items():
        if data:
            sents.append(str(val))
            #TRACE: If there is something wrong with inferVector, or val, then targ_vecs will be all 0
            # val comes from target_matches
            targ_vecs.append(par_vec.inferVector(str(val))) 
            targs[str(val)] = key
        else:
            for line in val:
                #print(line)
                sents.append(str(line))
                targ_vecs.append(par_vec.inferVector(str(line)))
                targs[str(line)] = key
    return targs, targ_vecs, sents


def convertPdf(document_conversion, config, file):
    '''
    Convert a pdf file into a txt file 

    Args:
        document_conversion (DocumentConversionV1) : Instance of the Document Conversion service.
        config (dict)                              : A config object that defines tags and structure 
                                                     in the conversion output.
        file (string)                              : path/filename to be converted.

    Returns:
        void: No return value. Txt file will be saved to the same directory
              as code.
    '''    
    with open(file, 'rb') as pdf_file:
        try:
            response = document_conversion.convert_document(document=pdf_file, config=config)
            document_text = response.text
            
            with open(file[:-4]+'.txt', 'w')  as f:
                f.write(document_text)
        
        except Exception as e:
            logger.error(f"Failed to convert: {file}")
            logger.error(e)
            
            



def get_all_matches(data, par_vec: CustomParVec, sents, targ_vecs, targs, num_out):
    sdg_matches = []
    targ_matches = []
    total_icaad_sdg = 0
    total_icaad_targets = 0
    score_dict = {}
    
    for series in data.iterrows():
        idx, row = series[0], series[1]
        line = row['RECOMMENDATION']

        if len(line) > 0:
            top_matches = par_vec.getMostSimilar(line, num_out, .1, sents, targ_vecs)
            
            icaad_sdgs = row['SDGs'].split(',')

            if '' in icaad_sdgs:
                icaad_sdgs.remove('')
            icaad_targets = row['Subcategory'].split(',')
            
            if '' in icaad_targets:
                icaad_targets.remove('')
            total_icaad_sdg += len(icaad_sdgs)
            total_icaad_targets += len(icaad_targets)
            
            for match in top_matches:
                key = targs[match[1]]
                sdg = key.split('.')
                
                if line in score_dict:
                    score_dict[line].append(key)
                else:
                    score_dict[line] = [key]
                
                if sdg[0] in icaad_sdgs:
                    sdg_matches.append(sdg[0])
                    icaad_sdgs.remove(sdg[0])
                if key in icaad_targets:
                    targ_matches.append(key)
                    icaad_targets.remove(key)
                    
    return total_icaad_sdg, total_icaad_targets, sdg_matches, targ_matches, score_dict



#Functions to conduct a RIA


#We will use our custom embedding technique to infer vectors for each of our ground truth sentences. 
#Then, given the policy documents for a country that we wish to produce the RIA for, we will compare the 
#similarity of each sentence/paragraph with the sentences from the ground truth. 
#Those sentences with the highest cosine similarity will be marked as matching the same target as the ground truth.

def ria(documents_path, policy_documents, model: CustomParVec, sents, targ_vecs, targs):
    '''
    Find the sentences/paragaraphs of policy documents that most match each target
    
    Args:
        documents_path (string)      : Directory holding all documents.
        policy_documents (list[str]) : List of policy documents for country RIA is to be conducted for.
        model (CustomParVec)         : Embedding model to be used.
        sents (list[str])            : list of ground truth sentences to enhance semantic searching.
        targ_vecs (list[np.array])   : list of vector embeddings for those ground truth sentences.
        targs (dict)                 : Dictionary of sentence to target

    Returns:
        score_dict (dict) : dictionary of target to ordered sentences found that match the target
    '''
    score_dict = {}
    for policy_document in policy_documents:
        with open(os.path.join(documents_path, policy_document), encoding="utf-8") as file:
            for line in file:
                if len(line) > 30:
                    top_matches = model.getMostSimilar(line, 125, 0.01, sents, targ_vecs)
                    for match in top_matches:
                        key = targs[match[1]]
                        if key in score_dict:
                            score_dict[key].add((match[0], line))
                        else:
                            score_dict[key] = set({(match[0], line)})
    return score_dict


def riaPDF(documents_path, policy_documents, model, sents, targ_vecs, targs):
    '''
    Find the sentences/paragaraphs of policy documents that most match each target.
    
    Args:
        documents_path (string)      : Directory holding all documents.
        policy_documents (list[str]) : List of policy documents for country RIA is to be conducted for.
        model (CustomParVec)         : Embedding model to be used.
        sents (list[str])            : list of ground truth sentences to enhance semantic searching.
        targ_vecs (list[np.array])   : list of vector embeddings for those ground truth sentences.
        targs (dict)                 : Dictionary of sentence to target

    Returns:
        score_dict (dict) : dictionary of target to ordered sentences found that match the target
    '''
    score_dict = {}
    for policy_document in policy_documents:
        try:
            inputpdf = PdfFileReader(os.path.join(documents_path, policy_document), "rb")
        except Exception as e:
            logger.error(f"Failed to open: {policy_document}")
            logger.error(e)
            continue

        for i in range(inputpdf.numPages):
            output = PdfFileWriter()
            output.addPage(inputpdf.getPage(i))
            newname = policy_document[:-4] + "-" + str(i+1) + '.pdf'
            outputStream = open(newname, "wb")
            output.write(outputStream)
            outputStream.close()

#NEEDS TO BE COMMENTED OUT WHEN COVERTPDF IS RE_WRITTEN            convertPdf(document_conversion, config, newname)

            try:
                with open(newname[:-4]+'.txt') as file:
                    for line in file:
                        if len(line) > 30:
                            
                            top_matches = model.getMostSimilar(line, 125, 0.01, sents, targ_vecs)
                            for match in top_matches:
                                key = targs[match[1]]
            
                                if key in score_dict:
                                    score_dict[key].add((match[0], line, policy_document, i+1))
                                else:
                                    score_dict[key] = set({(match[0], line, policy_document, i+1)})
            
            except Exception as e:
                logger.error(f"Failed to open: {newname[:-4]+'.txt'}")
                logger.error(e)
                continue
            
            os.remove(newname)
            os.remove(newname[:-4]+'.txt')
    return score_dict



#Functions to view RIA Results
def get_matches2(target, target_dict, num_matches=1000):
    '''
    Returns the specified number of matches of a target along with its document of
    origin and page number in a target dictionary ordered by cosine similarity 
    
    Args:
        target (string)    : Target to return matches for
        target_dict (dict) : Dictionary of target matches
        num_matches (int)  : Number of matches to be returned.
   

    Returns:
        list(Page Number, Document, Text) : List of num_matches sentences/paragraphs that correspond to the 
                                            specified target.
    '''
    ordered = [['', '', item[1]] for item in sorted(target_dict[target], reverse = True)]
    # PDF CHANGE: Once PDF reader methods have been implemented, replace above line by one below
    #ordered = [[item[3], item[2], item[1]] for item in sorted(target_dict[target], reverse = True)]
    return ordered[:num_matches]


def get_matches(target, target_dict, num_matches=1000):
    '''
    Returns the specified number of matches of a tagret in a target dictionary ordered by cosine similarity 
    
    Args:
        target (string)    : Target to return matches for
        target_dict (dict) : Dictionary of target matches
        num_matches (int)  : Number of matches to be returned.
   

    Returns:
        (list) : List of num_matches sentences/paragraphs that correspond to the specified target.
    '''
    ordered = [item[1] for item in sorted(target_dict[target], reverse = True)]
    return list(reversed(OrderedDict.fromkeys(reversed(ordered))))[:num_matches]

def lookup_matches(target, target_dict):
    results = get_matches(target, target_dict)
    for result in results:
        print(result)


#Functions to evaluate RIA Results
def evaluateByTarget(score_dict, test_target_matches, num):
    '''
    Finds matches with prior RIA as the number of sentences outputted increases
    
    Args:
        score_dict (dict)          : Our target matches
        test_target_matches (dict) : Target matches from prior RIA
        num (int)                  : Number of output sentences to match.
   

    Returns:
        (dict) : Dictionary of how many matches were found after each sentence per target
    '''
    truths = []
    match_by_sent = {}
    truth_dict = {}
    check = []

    for target in score_dict.keys():
        for result in get_matches(target, score_dict, num): 
            if target in test_target_matches and len(test_target_matches[target]) > 1:
                sentences = result.split('.')
                for sent in sentences:
                    for ground_truth in test_target_matches[target]:
                        score = sm(None, ground_truth, sent).ratio()
                        if score > 0.50:
                            if score < .55:
                                check.append((ground_truth, sent))
                            if target in truth_dict and ground_truth not in truths:
                                truths.append(ground_truth)
                                truth_dict[target].append(ground_truth)
                            elif target not in truth_dict and ground_truth not in truths:
                                truth_dict[target] = [ground_truth]
                                truths.append(ground_truth)

                if target in truth_dict:
                    if target in match_by_sent:
                        match_by_sent[target].append(len(truth_dict[target])/(len(test_target_matches[target])-1))
                    else:
                        match_by_sent[target] = [len(truth_dict[target])/(len(test_target_matches[target])-1)]
                else:
                    if target in match_by_sent:
                        match_by_sent[target].append(0)
                    else:
                        match_by_sent[target] = [0]
    return match_by_sent


def avgMatches(match_by_sent, test_target_matches, num):
    '''
    Finds the average percent matches with prior RIA for all targets as the number of sentences outputted increases
    
    Args:
        match_by_sent (dict)       : Dictionary of percent matches by target per sentence
        test_target_matches (dict) : Target matches from prior RIA
        num (int)                  : Number of output sentences to match.
   

    Returns:
        (dict) : Dictionary of how many matches were found after each sentence per target
    '''
    avg_new = []
    for i in range(num):
        adder, counter = 0, 0
        for key in match_by_sent:
            try:
                adder += (match_by_sent[key][i] * (len(test_target_matches[key])-1))
                counter += (len(test_target_matches[key])-1)
            except Exception as e:
                logger.warning(e) 
                adder += (match_by_sent[key][-1] * (len(test_target_matches[key])-1))
                counter += (len(test_target_matches[key])-1)
        
        avg = adder/counter if counter != 0 else None
        avg_new.append(avg)
    return avg_new


#Functions to generate the final output
def getResults(score_dict, num_matches):
    '''
    Retrieve the items wanted from a score_dict
    
    Args:
        score_dict (dict)  : target/result dictionary.
        num_matches (int)  : Number of matches to be returned per target
   
    '''
    results = {}
    for key in score_dict:
        results[key] = get_matches2(key, score_dict, num_matches) #was get_matches2(key, score_dict, num_matches) for pdf doc pages

    return results


import pandas as pd
import xlsxwriter
from openpyxl import load_workbook

def generateSpreadsheet(results, name):
    '''
    Generate an excel spreadhseet of the results in which each sheet corresponds to a target.
    The first columns is left blank for evaluation, second column is the page number, third the origin document, and
    fourth, the target match.
    
    Args:
        results (dict) : target/result dictionary.
        name (string)  : Spreadsheet name.
   
    '''
    # Create a Pandas dataframe from the data.
    df = pd.DataFrame.from_dict(results, orient='index')
    x = df.transpose()

    xbook = xlsxwriter.Workbook(name)
    #header = ['Page #', 'Policy Document', 'Target Match']
    
    # Convert the dataframe to an XlsxWriter Excel object.
    for col in sorted(x, key=lambda x: float(x)):
        xsheet = xbook.add_worksheet(str(col))
        #xsheet.write_row(0, 1, header)
        for i in range(len(x[col])):
            xsheet.write_row(i, 1, x[col].loc[i])
            
            
    xbook.close()
    
    
#Functions to update the model for the next RIA
def getUpdates(excel_results):
    '''
    Gets the sentences that were properly matched to a target evaluated by policy experts from a RIA conducted.
    
    Args:
        excel_results (string): Excel workbook with evaluated results.
    
    Returns:
        updates(dict) : Dictionary of target to sentences that were properly matches as evaluated by policy experts.
    '''
    updates = {}
    wb = load_workbook(excel_results)
    for sheet in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(sheet)
        updates[sheet] = []
        for i in range(1, 6):
            if ws.cell(row = i, column = 1).value == 1:
                updates[sheet].append(ws.cell(row = i, column = 4).value)
                
    return updates


def updateGroundTruth(new_truths):
    '''
    Updates the model with the new truth from the most recent RIA 
    
    Args:
        new_truths (dict): Dictionary of new sentence matches(value:list) for the target(key)

    '''
    try:
        shelf = shelve.open('RIA_Data')
        ground_truth = shelf['ground_truth']
        shelf.close()
    except Exception as e:
        logger.error(e)
        shelf.close()
    
    for key in new_truths:
        doc = ''
        for val in new_truths[key]:
            if type(val) == str:
                doc += val
                doc += ' '
        ground_truth[key][0] += doc
    
    try:
        shelf = shelve.open('RIA_Data')
        shelf['ground_truth'] = ground_truth
        shelf.close()
    except Exception as e:
        logger.error(e)
        shelf.close()


def open_shelve_object(shelf_name: str, obj_name: str):
    if os.path.exists(shelf_name + '.db'):
        shelf = shelve.open(shelf_name)
        obj = shelf[obj_name]
        shelf.close()
        return obj
    else:
        logger.warning(f"Shelf {shelf_name} does not exist")
        return None


def create_shelve_object(shelf_name: str, source_data: Any, obj_name: str):

    shelf_path = Path(shelf_name)
    
    if not os.path.exists(shelf_path.parent):
        os.makedirs(shelf_path.parent)

    shelf = shelve.open(shelf_name)
    shelf[obj_name] = source_data
    shelf.close()